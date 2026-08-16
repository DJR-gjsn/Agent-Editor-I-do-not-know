"""
全组件连线测试 — 通过 /api/chat 逐一测试，生成 Excel 报告
"""
import json, time, requests, sys, io, os
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

API = "http://localhost:5000/api/chat"
API_KEY = ""
API_BASE = "https://api.deepseek.com/v1"
MODEL = "deepseek-chat"

# Get all tools and project
all_tools = requests.get("http://localhost:5000/api/tools/definitions").json()
tool_index = {t["function"]["name"]: t for t in all_tools}

with open('data/projects/proj_b3f7d16fb5.json', 'r', encoding='utf-8') as f:
    proj = json.load(f)
comps = {c['id']: c for c in proj['layout']['components']}
conns = proj['layout']['connections']

# Component type → tool names mapping (from app.js TOOL_NAME_MAP)
comp_tools = {
    'web_search': ['web_search'],
    'calculator': ['calculator'],
    'time_query': ['get_current_time'],
    'url_fetch': ['url_fetch'],
    'code_executor': ['code_executor'],
    'text_tools': ['text_analyze', 'text_format'],
    'mcp_word': ['word_create', 'word_add_heading', 'word_add_paragraph', 'word_save'],
    'mcp_excel': ['excel_create', 'excel_write_cell', 'excel_save'],
    'mcp_ppt': ['ppt_create', 'ppt_add_slide', 'ppt_save'],
    'mcp_pdf': ['pdf_create'],
    'file_ops': ['file_write', 'file_read'],
    'json_query': ['json_query'],
    'mcp_encoding': ['hash_compute', 'base64_encode'],
    'mcp_translate': ['translate_text', 'detect_language'],
    'mcp_weather': ['weather_current'],
    'mcp_geocode': ['ip_geolocation'],
    'mcp_system': ['system_info'],
    'mcp_clipboard': ['clipboard_read'],
    'memory_summarizer': ['memory_summarize'],
    'plan': ['plan_generate'],
    'executor': ['executor_run'],
    'mcp_finance': ['currency_convert'],
    'mcp_calendar': ['calendar_list'],
    'mcp_database': ['db_list_tables'],
    'mcp_git': ['git_status'],
    'mcp_navigation': ['nav_search_place'],
    'mcp_email': ['email_send'],
    'vector_memory': ['embeddings_index', 'embeddings_search'],
}

# Test prompts
test_prompts = {
    'web_search': "search Python 3.14 latest version",
    'calculator': "calculate 123*456",
    'time_query': "what time is it now",
    'url_fetch': "fetch https://httpbin.org/json",
    'code_executor': "use Python to compute sum(range(1,101))",
    'text_tools': "analyze word count of: hello world test",
    'mcp_word': "create a Word doc titled TestReport",
    'mcp_excel': "create Excel, write 'TestData' in cell A1",
    'mcp_ppt': "create a PPT titled TestPresentation",
    'mcp_pdf': "convert 'Hello World' to PDF",
    'file_ops': "create file test_auto.txt with content 'auto test'",
    'json_query': "extract field 'a' from JSON: {\"a\":1}",
    'mcp_encoding': "compute MD5 hash of 'hello'",
    'mcp_translate': "detect language of 'Hello World'",
    'mcp_weather': "check Beijing weather",
    'mcp_geocode': "lookup IP 8.8.8.8 location",
    'mcp_system': "get system info",
    'mcp_clipboard': "read clipboard",
    'memory_summarizer': "summarize this conversation: user asked about sort algorithms, AI explained bubble sort",
    'plan': "make a 2-step plan for writing a test report",
    'executor': "execute: step1 calculate 1+1",
    'mcp_finance': "convert 100 USD to CNY",
    'mcp_calendar': "list today's calendar events",
    'mcp_database': "list database tables",
    'mcp_git': "check git status",
    'mcp_navigation': "search nearby restaurants",
    'mcp_email': "send test email to test@example.com",
    'vector_memory': "add this document to the vector knowledge base: Python is a programming language for data science",
}

# Meta components (no tools, modify behavior)
meta_types = {
    'working_memory': '注入临时工作记忆到 system prompt',
    'system_prompt': '注入自定义 system prompt',
    'vision': '图像识别（需配合 vision API）',
    'function_calling': '注册自定义 Function Calling Schema',
    'json_mode': '强制 LLM 输出 JSON 格式',
    'sequential_executor': '约束 LLM 按顺序调用工具',
    'loop': 'Plan→Execute→Reflect 循环控制',
    'agent': 'Plan→Execute→Reflect 编排中枢',
    'skills_manager': '技能管理中枢（聚合 skill 组件）',
    'reflection': '反思评估（评估执行结果）',
    'memory': '对话记忆持久化存储',
}

results = []

for comp_id, comp in sorted(comps.items()):
    ctype = comp['type']
    if ctype == 'llm':
        continue

    # Meta component
    if ctype in meta_types:
        results.append({
            'id': comp_id, 'type': ctype, 'category': '编排/增强',
            'status': 'PASS', 'elapsed': '-',
            'detail': '无需工具调用: ' + meta_types[ctype],
            'tools_called': '-', 'issues': ''
        })
        continue

    # Tool component
    tool_names = comp_tools.get(ctype, [])
    if not tool_names:
        results.append({
            'id': comp_id, 'type': ctype, 'category': '未知',
            'status': 'SKIP', 'elapsed': '-',
            'detail': '未配置工具映射',
            'tools_called': '-', 'issues': '未知组件类型'
        })
        continue

    test_tools = [{"type": "function", "function": tool_index[t]["function"]}
                  for t in tool_names if t in tool_index]
    if not test_tools:
        results.append({
            'id': comp_id, 'type': ctype, 'category': '工具',
            'status': 'SKIP', 'elapsed': '-',
            'detail': '工具未在后端注册',
            'tools_called': '-', 'issues': '工具注册缺失'
        })
        continue

    prompt = test_prompts.get(ctype, f"use {tool_names[0]} tool")
    payload = {
        "messages": [{"role": "user", "content": prompt}],
        "model": MODEL, "api_base": API_BASE, "api_key": API_KEY,
        "tools": test_tools, "max_tool_rounds": 5,
        "max_tokens": 1024, "temperature": 0.1, "stream": False
    }

    print(f"Testing [{comp_id}] {ctype}...", end=' ', flush=True)

    try:
        t0 = time.time()
        resp = requests.post(API, json=payload, timeout=60)
        elapsed = round(time.time() - t0, 1)

        calls = []; tool_results = []; errors = []; content = ""
        for line in resp.text.split('\n'):
            if not line.startswith('data: '): continue
            ds = line[6:]
            if ds == '[DONE]': break
            try: d = json.loads(ds)
            except: continue
            if d.get('error'): errors.append(d['error'])
            if d.get('tool_calls'):
                for tc in d['tool_calls']: calls.append(tc['name'])
            if d.get('tool_result'):
                tr = d['tool_result']
                tool_results.append({"name": tr['name'], "result": tr['result'][:300]})
            if d.get('choices'):
                c = d['choices'][0].get('delta',{}).get('content','')
                if c: content += c

        # Determine status
        status = "PASS"
        detail = ""
        issues = ""

        if errors:
            status = "FAIL"; detail = errors[0][:150]; issues = errors[0][:200]
        elif not calls:
            status = "FAIL"; detail = "LLM未调用工具: " + content[:60]
            issues = "LLM拒绝或未识别工具调用"
        else:
            for tr in tool_results:
                r = tr["result"]
                for kw in ['Error:', '连接失败', 'Traceback', 'AttributeError',
                           'ModuleNotFoundError', 'API_KEY', 'not configured',
                           'latin-1', 'UnicodeEncodeError', 'Connection refused',
                           'timed out', '不支持', '未配置', '失败']:
                    if kw in r:
                        status = "WARN"; detail = r[:150]; issues = r[:200]; break
                if status == "WARN":
                    break
            if status == "PASS":
                detail = "called: " + ", ".join(set(calls))

        results.append({
            'id': comp_id, 'type': ctype, 'category': '工具',
            'status': status, 'elapsed': elapsed,
            'detail': detail, 'tools_called': ", ".join(calls), 'issues': issues
        })
        print(f"[{status}] ({elapsed}s)")

    except Exception as e:
        results.append({
            'id': comp_id, 'type': ctype, 'category': '工具',
            'status': 'FAIL', 'elapsed': 0,
            'detail': str(e)[:150], 'tools_called': '', 'issues': str(e)[:200]
        })
        print(f"[FAIL] {e}")

# Sort and save
results.sort(key=lambda r: (r['status'] != 'PASS', r['type']))

with open("test_all_results.json", "w", encoding="utf-8") as f:
    json.dump(results, f, ensure_ascii=False, indent=2)

# Summary
pass_c = sum(1 for r in results if r['status'] == 'PASS')
warn_c = sum(1 for r in results if r['status'] == 'WARN')
fail_c = sum(1 for r in results if r['status'] == 'FAIL')
skip_c = sum(1 for r in results if r['status'] == 'SKIP')

print(f"\n{'='*60}")
print(f"RESULTS: {pass_c} PASS, {warn_c} WARN, {fail_c} FAIL, {skip_c} SKIP")
print(f"{'='*60}")
for r in results:
    if r['status'] != 'PASS':
        icon = 'WARN' if r['status'] == 'WARN' else ('FAIL' if r['status'] == 'FAIL' else 'SKIP')
        print(f"  [{icon}] {r['type']:25s} | {r['detail'][:100]}")

# Generate Excel
print(f"\nGenerating Excel...")
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Component Status"

    # Styles
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    warn_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    skip_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # Headers
    headers = ['#', '组件类型', 'ID', '分类', '状态', '耗时(s)', '工具调用', '详情', '问题']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    # Data
    for i, r in enumerate(results, 2):
        ws.cell(row=i, column=1, value=i-1).border = thin_border
        ws.cell(row=i, column=2, value=r['type']).border = thin_border
        ws.cell(row=i, column=3, value=r['id']).border = thin_border
        ws.cell(row=i, column=4, value=r['category']).border = thin_border

        status_cell = ws.cell(row=i, column=5, value=r['status'])
        status_cell.border = thin_border
        status_cell.alignment = Alignment(horizontal='center')
        if r['status'] == 'PASS': status_cell.fill = pass_fill
        elif r['status'] == 'WARN': status_cell.fill = warn_fill
        elif r['status'] == 'FAIL': status_cell.fill = fail_fill
        else: status_cell.fill = skip_fill

        ws.cell(row=i, column=6, value=r['elapsed']).border = thin_border
        ws.cell(row=i, column=7, value=r['tools_called']).border = thin_border
        ws.cell(row=i, column=8, value=r['detail']).border = thin_border
        ws.cell(row=i, column=9, value=r['issues']).border = thin_border

        for col in range(1, 10):
            ws.cell(row=i, column=col).alignment = Alignment(vertical='center', wrap_text=(col >= 7))

    # Summary row
    summary_row = len(results) + 3
    ws.cell(row=summary_row, column=1, value="汇总").font = Font(bold=True, size=12)
    ws.merge_cells(start_row=summary_row, start_column=1, end_row=summary_row, end_column=2)

    summary_data = [
        ('PASS', pass_c, pass_fill),
        ('WARN', warn_c, warn_fill),
        ('FAIL', fail_c, fail_fill),
        ('SKIP', skip_c, skip_fill),
        ('TOTAL', len(results), None),
    ]
    for j, (label, count, fill) in enumerate(summary_data):
        row = summary_row + 1 + j
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row, column=2, value=count).font = Font(bold=True)
        if fill:
            ws.cell(row=row, column=1).fill = fill

    # Column widths
    widths = [5, 22, 6, 12, 8, 10, 35, 55, 50]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    # Freeze header
    ws.freeze_panes = 'A2'

    output_path = 'component_status_report.xlsx'
    wb.save(output_path)
    print(f"Excel saved: {output_path}")

except ImportError:
    print("openpyxl not installed, installing...")
    os.system("pip install openpyxl -q")
    print("Please re-run the script to generate Excel.")

print("Done!")
