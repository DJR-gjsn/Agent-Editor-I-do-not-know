import json, time, requests, sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

API = "http://localhost:5000/api/chat"
API_KEY = ""
API_BASE = "https://api.deepseek.com/v1"
MODEL = "deepseek-chat"

skills = requests.get("http://localhost:5000/api/skills").json()

test_prompts = {
    "document": "I need to create a Word document for meeting notes",
    "frontend-design": "Design a login page UI for a banking app",
    "ui-ux-pro-max": "Design UX for an e-commerce checkout page",
    "find-skills": "Find skills related to code testing",
    "skill-creator": "I want to create a new API doc generator skill",
    "superpowers": "Analyze this code quality: def add(a,b): return a+b",
    "pua": "Give me execution advice for my new project",
}

skill_results = []

for skill_info in skills:
    sid = skill_info['id']
    prompt = test_prompts.get(sid, "Use skill: " + sid)

    payload = {
        "messages": [{"role": "user", "content": prompt}],
        "model": MODEL, "api_base": API_BASE, "api_key": API_KEY,
        "smart_mode": True,
        "available_skills": [{"id": sid, "name": skill_info['name']}],
        "max_tool_rounds": 5, "max_tokens": 2048, "temperature": 0.3, "stream": False
    }

    print("Testing: " + sid + "...", end=' ', flush=True)

    try:
        t0 = time.time()
        resp = requests.post(API, json=payload, timeout=120)
        elapsed = round(time.time() - t0, 1)

        calls = []; tool_results = []; errors = []; content = ""
        for line in resp.text.split('\n'):
            if not line.startswith('data: '): continue
            ds = line[6:]
            if ds == '[DONE]': break
            try: d = json.loads(ds)
            except: continue
            if d.get('error'): errors.append(d['error'])
            if d.get('tool_calls'):
                for tc in d['tool_calls']: calls.append(tc['name'])
            if d.get('tool_result'):
                tr = d['tool_result']
                tool_results.append({"name": tr['name'], "result": tr['result'][:300]})
            if d.get('choices'):
                c = d['choices'][0].get('delta',{}).get('content','')
                if c: content += c

        use_skill_called = "use_skill" in calls
        skill_prompt_len = 0
        skill_ok = False
        for tr in tool_results:
            if tr['name'] == 'use_skill':
                r = tr['result']
                skill_prompt_len = len(r)
                if 'not found' not in r.lower() and r:
                    skill_ok = True

        status = "PASS"
        detail = ""
        issues = ""

        if errors:
            status = "FAIL"; detail = errors[0][:120]; issues = errors[0][:200]
        elif use_skill_called and skill_ok:
            detail = "Skill activated ({} chars), calls: {}".format(skill_prompt_len, calls)
        elif use_skill_called and not skill_ok:
            status = "FAIL"; detail = "use_skill returned empty/not-found"
            issues = "_smart_skills_cache not populated"
        elif not use_skill_called and content:
            detail = "LLM handled directly: " + content[:100]
        else:
            status = "FAIL"; detail = "No tool calls and no content"
            issues = "Unknown failure"

        skill_results.append({
            'type': 'skill_' + sid, 'id': 'skill-' + sid, 'category': 'Skills system',
            'status': status, 'elapsed': elapsed,
            'detail': detail, 'tools_called': ", ".join(calls), 'issues': issues
        })
        print("[{}] ({}s) {}".format(status, elapsed, detail[:80]))

    except Exception as e:
        skill_results.append({
            'type': 'skill_' + sid, 'id': 'skill-' + sid, 'category': 'Skills system',
            'status': 'FAIL', 'elapsed': 0,
            'detail': str(e)[:150], 'tools_called': '', 'issues': str(e)[:200]
        })
        print("[FAIL] " + str(e))

# Save and merge
with open("test_skills_detailed.json", "w", encoding="utf-8") as f:
    json.dump(skill_results, f, ensure_ascii=False, indent=2)

with open("test_all_results.json", "r", encoding="utf-8") as f:
    all_results = json.load(f)

# Remove old skill entries if any, add new
all_results = [r for r in all_results if not r['type'].startswith('skill_')]
all_results.extend(skill_results)

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

all_results.sort(key=lambda r: ({'PASS':0,'WARN':1,'FAIL':2,'SKIP':3}[r['status']], r['type']))

wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Component Status'

hdr_font = Font(bold=True, size=11, color='FFFFFF')
hdr_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
pass_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
warn_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
fail_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
skip_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

hdrs = ['#', 'Component Type', 'ID', 'Category', 'Status', 'Time(s)', 'Tools Called', 'Detail', 'Issues']
for col, h in enumerate(hdrs, 1):
    c = ws.cell(row=1, column=col, value=h)
    c.font = hdr_font; c.fill = hdr_fill
    c.alignment = Alignment(horizontal='center', vertical='center'); c.border = border

fills = {'PASS': pass_fill, 'WARN': warn_fill, 'FAIL': fail_fill, 'SKIP': skip_fill}
for i, r in enumerate(all_results, 2):
    ws.cell(row=i, column=1, value=i-1).border = border
    ws.cell(row=i, column=2, value=r['type']).border = border
    ws.cell(row=i, column=3, value=str(r['id'])).border = border
    ws.cell(row=i, column=4, value=r['category']).border = border
    sc = ws.cell(row=i, column=5, value=r['status'])
    sc.border = border; sc.alignment = Alignment(horizontal='center')
    if r['status'] in fills: sc.fill = fills[r['status']]
    ws.cell(row=i, column=6, value=str(r['elapsed']) if r['elapsed'] != '-' else '-').border = border
    ws.cell(row=i, column=7, value=str(r['tools_called'])).border = border
    ws.cell(row=i, column=8, value=str(r['detail'])).border = border
    ws.cell(row=i, column=9, value=str(r['issues'])).border = border
    for col in range(1, 10):
        ws.cell(row=i, column=col).alignment = Alignment(vertical='center', wrap_text=(col >= 7))

pc = sum(1 for r in all_results if r['status']=='PASS')
wc = sum(1 for r in all_results if r['status']=='WARN')
fc = sum(1 for r in all_results if r['status']=='FAIL')
sc = sum(1 for r in all_results if r['status']=='SKIP')

sr = len(all_results) + 3
ws.cell(row=sr, column=1, value='Summary').font = Font(bold=True, size=12)
ws.merge_cells(start_row=sr, start_column=1, end_row=sr, end_column=2)
for j, (lbl, ct, fl) in enumerate([('PASS', pc, pass_fill), ('WARN', wc, warn_fill), ('FAIL', fc, fail_fill), ('SKIP', sc, skip_fill), ('TOTAL', len(all_results), None)]):
    row = sr + 1 + j
    ws.cell(row=row, column=1, value=lbl).font = Font(bold=True)
    ws.cell(row=row, column=2, value=ct).font = Font(bold=True)
    if fl: ws.cell(row=row, column=1).fill = fl

widths = [5, 26, 12, 16, 8, 10, 42, 64, 56]
for col, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(col)].width = w
ws.freeze_panes = 'A2'

outpath = 'D:/myxiangfa-MCPxuexi/component_status_report.xlsx'
wb.save(outpath)

# Save updated JSON
with open("test_all_results.json", "w", encoding="utf-8") as f:
    json.dump(all_results, f, ensure_ascii=False, indent=2)

print("\nExcel updated: " + outpath)
print("Skills added: {}".format(len(skill_results)))
print("Total in Excel: {} rows".format(len(all_results)))
print("PASS={} WARN={} FAIL={} SKIP={}".format(pc, wc, fc, sc))
