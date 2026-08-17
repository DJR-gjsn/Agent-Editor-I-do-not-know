"""生成组件功能清单 Excel：从 static/app.js 提取所有组件定义"""
import json
import re

SRC = "static/app.js"
OUT = "组件功能清单.xlsx"


def extract_object(name, src):
    """提取 const NAME = { ... }; 的文本块（括号配平）"""
    start = src.index("const %s = {" % name)
    i = src.index("{", start)
    depth = 0
    j = i
    while j < len(src):
        if src[j] == "{":
            depth += 1
        elif src[j] == "}":
            depth -= 1
            if depth == 0:
                return src[i : j + 1]
        j += 1
    return None


def parse_defs(block):
    """解析 COMPONENT_DEFS：key -> {title, description}"""
    defs = {}
    pos = 0
    while True:
        m = re.search(r"(\w+): \{", block[pos:])
        if not m:
            break
        key = m.group(1)
        start = pos + m.end() - 1
        depth = 1
        j = start + 1
        while j < len(block) and depth > 0:
            if block[j] == "{":
                depth += 1
            elif block[j] == "}":
                depth -= 1
            j += 1
        body = block[start:j]
        t = re.search(r"title:\s*'([^']*)'", body)
        d = re.search(r"description:\s*'([^']*)'", body, re.S)
        defs[key] = {
            "title": t.group(1) if t else key,
            "description": (d.group(1).replace("\\'", "'") if d else ""),
        }
        pos = j
    return defs


def parse_map(block):
    """解析 KEY: ['a','b'] 映射"""
    out = {}
    for m in re.finditer(r"(\w+):\s*\[([^\]]*)\]", block):
        items = re.findall(r"'([^']*)'", m.group(2))
        out[m.group(1)] = items
    return out


src = open(SRC, encoding="utf-8").read()
defs = parse_defs(extract_object("COMPONENT_DEFS", src))
cats = parse_map(extract_object("COMPONENT_CATEGORIES", src))
tools = parse_map(extract_object("TOOL_NAME_MAP", src))

print(f"组件总数: {len(defs)}")

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "组件功能清单"

header_font = Font(bold=True, size=11, color="FFFFFF")
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
thin = Border(left=Side(style="thin"), right=Side(style="thin"),
              top=Side(style="thin"), bottom=Side(style="thin"))

headers = ["#", "组件类型", "名称", "分类", "工具", "功能说明"]
for col, h in enumerate(headers, 1):
    c = ws.cell(row=1, column=col, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = thin

cat_color = {"核心": "D9E2F3", "编排": "E2EFDA", "流程": "FFF2CC",
             "记忆": "FCE4D6", "Skill": "E4DFEC", "Tools": "DDEBF7",
             "MCP": "EDEDED"}
row = 2
for key in sorted(defs.keys()):
    d = defs[key]
    cat = (cats.get(key) or ["其他"])[0]
    tl = ", ".join(tools.get(key, [])) if tools.get(key) else "—"
    ws.cell(row=row, column=1, value=row - 1).border = thin
    ws.cell(row=row, column=2, value=key).border = thin
    ws.cell(row=row, column=2).font = Font(name="Consolas", size=10)
    ws.cell(row=row, column=3, value=d["title"]).border = thin
    c4 = ws.cell(row=row, column=4, value=cat)
    c4.border = thin
    c4.fill = PatternFill(start_color=cat_color.get(cat, "F2F2F2"),
                          end_color=cat_color.get(cat, "F2F2F2"), fill_type="solid")
    ws.cell(row=row, column=5, value=tl).border = thin
    ws.cell(row=row, column=6, value=d["description"]).border = thin
    ws.cell(row=row, column=6).alignment = Alignment(wrap_text=True, vertical="center")
    row += 1

widths = [5, 24, 16, 10, 46, 70]
for col, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(col)].width = w
ws.freeze_panes = "A2"

# 汇总
summary = row + 1
ws.cell(row=summary, column=1, value="汇总").font = Font(bold=True, size=12)
ws.merge_cells(start_row=summary, start_column=1, end_row=summary, end_column=2)
from collections import Counter
cnt = Counter((cats.get(k) or ["其他"])[0] for k in defs)
for j, (cat, n) in enumerate(sorted(cnt.items())):
    r = summary + 1 + j
    ws.cell(row=r, column=1, value=cat).font = Font(bold=True)
    ws.cell(row=r, column=2, value=f"{n} 个")
ws.cell(row=summary + len(cnt) + 1, column=1, value="TOTAL").font = Font(bold=True)
ws.cell(row=summary + len(cnt) + 1, column=2, value=f"{len(defs)} 个").font = Font(bold=True)

wb.save(OUT)
print(f"已生成: {OUT}（{len(defs)} 个组件）")
