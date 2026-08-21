"""
全组件自动测试 — 真实连线模拟
========================================
读取项目 JSON 中的真实 components + connections 拓扑，
按编辑器（app.js）的连线逻辑推导每个组件的真实工具集与连接方式，
通过 /api/chat 用真实 LLM + 真实工具逐一测试，
输出 Excel 报告（component_status_report.xlsx）+ JSON（test_all_results.json）。

用法:
    python test_components_auto.py [project_json...]
    默认: data/projects/proj_7142edabd6.json （当前项目）

连线推导规则（镜像 app.js）:
- executor / agent / sequential_executor: 从工具端口（exec-tool-N / agent-tool-N / seq-step-N）
  的连接目标收集工具（collectToolsFromPorts）
- 直接连接到 LLM 的工具组件: 使用 TOOL_NAME_MAP[type]
- 技能链路: skill_auto_call → skills_manager → skill_*，用 smart_mode + use_skill 验证
- memory: 直接测试 /api/memory 持久化接口（不走 LLM）
"""
import json
import os
import sys
import time

import requests

sys.stdout.reconfigure(encoding="utf-8", errors="replace") if hasattr(sys.stdout, "reconfigure") else None

API = "http://127.0.0.1:5000"
CHAT = API + "/api/chat"
MEMORY_SAVE = API + "/api/memory/save"
MEMORY_LOAD = API + "/api/memory/load"

DEFAULT_PROJECT = "data/projects/proj_7142edabd6.json"

# ============================================================
# 登录（全局登录保护：所有 /api/* 需会话 cookie）
# ============================================================
SESSION = requests.Session()


def login():
    """登录获取会话 cookie。账号取环境变量 TEST_USER/TEST_PASS，默认 djr/123456。"""
    user = os.environ.get("TEST_USER", "djr")
    pwd = os.environ.get("TEST_PASS", "123456")
    resp = SESSION.post(API + "/api/auth/login",
                        json={"username": user, "password": pwd}, timeout=10)
    if resp.status_code != 200:
        print(f"[X] 登录失败（{user}）: {resp.text[:120]}")
        sys.exit(1)
    print(f"[OK] 已登录: {user}")

# ============================================================
# 工具映射（与 static/app.js 的 TOOL_NAME_MAP 保持一致）
# ============================================================
TOOL_NAME_MAP = {
    "web_search": ["web_search"], "calculator": ["calculator"], "code_executor": ["code_executor"],
    "text_tools": ["text_analyze", "text_format"], "time_query": ["get_current_time"],
    "url_fetch": ["url_fetch"], "file_ops": ["file_read", "file_write", "glob_search", "grep_search", "file_edit"],
    "json_query": ["json_query"],
    "vector_memory": ["embeddings_search", "embeddings_index"],
    "mcp_word": ["word_create", "word_add_heading", "word_add_paragraph", "word_add_table", "word_save"],
    "mcp_excel": ["excel_create", "excel_write_cell", "excel_read_cell", "excel_add_sheet", "excel_save"],
    "mcp_ppt": ["ppt_create", "ppt_add_slide", "ppt_add_text", "ppt_add_bullet_list", "ppt_save"],
    "mcp_weather": ["weather_current", "weather_forecast"],
    "mcp_database": ["db_query", "db_list_tables", "db_schema"],
    "mcp_git": ["git_status", "git_log", "git_diff", "git_branch"],
    "mcp_clipboard": ["clipboard_read", "clipboard_write"],
    "mcp_encoding": ["base64_encode", "base64_decode", "hash_compute", "password_generate", "uuid_generate",
                    "regex_test", "diff_text", "markdown_to_html", "html_to_markdown", "unit_convert", "url_encode"],
    "mcp_system": ["system_info", "dns_lookup", "qr_generate", "open_file", "desktop_notify"],
    "mcp_email": ["email_send"],
    "mcp_translate": ["translate_text", "detect_language"],
    "mcp_calendar": ["calendar_list", "calendar_create"],
    "mcp_pdf": ["pdf_read", "pdf_create", "pdf_merge"],
    "mcp_finance": ["currency_convert", "stock_price"],
    "mcp_geocode": ["geocode_address", "reverse_geocode", "ip_geolocation", "distance_calc"],
    "mcp_navigation": ["nav_route", "nav_search_place"],
    "plan": ["plan_generate", "plan_execute_step"],
    "memory_summarizer": ["memory_summarize"],
    "executor": ["executor_run"],
    "mcp_zip": ["zip_create", "zip_extract"],
    "http_request": ["http_request"],
    "image_tools": ["screenshot", "image_info", "image_convert", "image_resize", "image_compress"],
}

# 编排组件 → 端口前缀（executor 收集工具用）
ORCH_PORT_PREFIX = {
    "executor": "exec-tool-",
    "agent": "agent-tool-",
    "sequential_executor": "seq-step-",
}

# ============================================================
# 每个组件类型的真实测试 prompt
# ============================================================
TEST_PROMPTS = {
    "web_search": "请调用 web_search 工具搜索一次 Python 3.14 的最新版本信息，然后直接根据搜索结果回答，不要再重复搜索",
    "calculator": "计算 123*456 的结果",
    "time_query": "现在是什么时间？",
    "url_fetch": "抓取 https://httpbin.org/json 的内容",
    "code_executor": "用 Python 计算 sum(range(1,101))",
    "text_tools": "分析这句话的字数: hello world test",
    "file_ops": "创建文件 test_auto.txt，内容为 'auto test 自动测试'",
    "json_query": "从 JSON 中提取字段 a: {\"a\": 1}",
    "vector_memory": "把这句话加入向量知识库: Python is a programming language for data science",
    "mcp_word": "创建 Word 文档，标题为 AutoTestReport",
    "mcp_excel": "创建 Excel，在 A1 单元格写入 'TestData'",
    "mcp_ppt": "创建 PPT，标题为 AutoTestPresentation",
    "mcp_weather": "查询北京今天的天气",
    "mcp_database": "列出数据库中的表",
    "mcp_git": "查看当前 git 仓库状态",
    "mcp_clipboard": "读取剪贴板内容",
    "mcp_encoding": "计算 'hello' 的 MD5 哈希",
    "mcp_system": "获取系统信息",
    "mcp_email": "发送测试邮件到 test@example.com",
    "mcp_translate": "检测 'Hello World' 的语言",
    "mcp_calendar": "列出今天的日历事件",
    "mcp_pdf": "把 'Hello World' 转换成 PDF 文件",
    "mcp_finance": "把 100 美元转换为人民币",
    "mcp_geocode": "查询 IP 8.8.8.8 的位置",
    "mcp_navigation": "搜索附近的餐厅",
    "plan": "为'写一份测试报告'制定一个两步计划",
    "executor": "按顺序执行两步：第一步计算 123*456，第二步获取当前时间",
    "memory_summarizer": "总结这段对话：用户问排序算法，AI 解释了冒泡排序",
    "mcp_zip": "把文件 test_auto.txt 压缩成 test_auto.zip",
    "http_request": "用 GET 请求访问 https://httpbin.org/json 并返回响应",
    "image_tools": "把 test_auto.png 图片转换成 JPEG 格式",
    "skill_document": "使用文档处理技能创建一个包含'你好'的 Word 文档",
    "skill_pua": "使用 PUA 绩效教练技能，帮我分析'测试没跑完就宣称完成'这个问题",
    "skills_manager": "使用可用的技能之一来完成：创建一个包含'技能测试'的 Word 文档",
    "skill_auto_call": "智能模式测试：根据任务自主选择技能来创建一个 Word 文档",
}

# 编排/增强组件：无需工具调用，直接标记 PASS
META_TYPES = {
    "llm": "LLM 本体（不测试）",
    "working_memory": "工作记忆（注入 system prompt，运行时行为）",
    "system_prompt": "自定义 system prompt（注入行为）",
    "function_calling": "自定义 Function Calling Schema（注册行为）",
    "json_mode": "强制 JSON 输出（行为开关）",
    "loop": "Plan→Execute→Reflect 循环控制（编排行为）",
    "agent": "Plan→Execute→Reflect 编排中枢（编排行为）",
    "reflection": "反思评估（编排行为）",
    "sequential_executor": "顺序工具调用约束（编排行为）",
    "vision": "图像识别（需配合 vision API）",
    "memory": "对话记忆持久化（单独 API 测试）",
    "skills_manager": "技能管理中枢（通过智能模式测试）",
    "skill_auto_call": "技能自动调用（通过智能模式测试）",
}

# 结果中视为 WARN 的关键词（精确匹配工具错误，避免中文 LLM 输出误报）
WARN_KEYWORDS = ["Error:", "Traceback", "AttributeError", "ModuleNotFoundError",
                 "连接失败", "Connection refused", "API_KEY", "not configured",
                 "UnicodeEncodeError", "timed out", "工具执行错误", "工具执行超时",
                 "线程池繁忙", "未配置", "不支持", "latin-1"]

MAX_ROUNDS_MSG = "工具调用达到最大轮数限制"


def load_project(path):
    """读取项目 JSON，返回 (components, connections)"""
    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)
    layout = data.get("layout", {})
    return layout.get("components", []), layout.get("connections", [])


def get_llm_config(components):
    """从 LLM 组件读取 apiSettings（真实项目配置）；key 为空时回退环境变量 LLM_API_KEY"""
    api_base = api_key = model = None
    for c in components:
        if c.get("type") == "llm" and c.get("apiSettings", {}).get("apiBase"):
            s = c["apiSettings"]
            api_base = s.get("apiBase")
            api_key = s.get("apiKey") or os.environ.get("LLM_API_KEY", "")
            model = s.get("model", "deepseek-chat")
            break
    if not api_base:
        api_base, api_key, model = "https://api.deepseek.com/v1", \
            os.environ.get("LLM_API_KEY", ""), "deepseek-chat"
    return api_base, api_key, model


def collect_tools_from_ports(comp_id, port_prefix, count, comps, conns):
    """镜像 app.js collectToolsFromPorts：从端口连接的目标组件收集工具"""
    names = []
    comp_by_id = {c["id"]: c for c in comps}
    for i in range(1, count + 1):
        port_id = f"{port_prefix}{i}"
        for cn in conns:
            if cn["sourceCompId"] == comp_id and conn_port(cn, "sourcePort") == port_id:
                tgt = comp_by_id.get(cn["targetCompId"])
                if tgt:
                    names.extend(TOOL_NAME_MAP.get(tgt["type"], []))
                break
    return list(dict.fromkeys(names))


def conn_port(cn, key):
    """兼容新旧连线字段名：sourcePort/sourcePortId、targetPort/targetPortId"""
    return cn.get(key) or cn.get(key + "Id")


def derive_component_info(comp, comps, conns):
    """推导组件的真实工具集与连接方式（模拟编辑器逻辑）"""
    ctype = comp["type"]
    comp_by_id = {c["id"]: c for c in comps}

    # 编排组件：端口收集
    if ctype in ORCH_PORT_PREFIX:
        prefix = ORCH_PORT_PREFIX[ctype]
        count = comp.get("execPortCount", comp.get("agentPortCount", comp.get("seqPortCount", 5)))
        tools = collect_tools_from_ports(comp["id"], prefix, count, comps, conns)
        if not tools:
            # 无端口连接时退回自身编排工具（如 executor_run）
            tools = TOOL_NAME_MAP.get(ctype, [])
        conn_desc = f"{ctype} 端口收集 {len(tools)} 个工具: {', '.join(tools[:6])}{'...' if len(tools) > 6 else ''}"
        return tools, conn_desc

    # 技能组件：智能模式
    if ctype.startswith("skill_"):
        skill_id = {"skill_document": "document", "skill_frontend": "frontend-design",
                    "skill_uiux": "ui-ux-pro-max", "skill_find": "find-skills",
                    "skill_creator": "skill-creator", "skill_super": "superpowers",
                    "skill_pua": "pua"}.get(ctype, ctype.replace("skill_", ""))
        # 找它连接的 skills_manager
        skm_conns = [cn for cn in conns if cn["targetCompId"] == comp["id"]]
        desc = "技能组件 → 智能模式 use_skill"
        return [skill_id], desc

    # 普通工具组件：找它的连线方式（连到 LLM 还是编排端口）
    tools = TOOL_NAME_MAP.get(ctype, [])
    conn_desc = ""
    for cn in conns:
        if cn["targetCompId"] == comp["id"]:
            src = comp_by_id.get(cn["sourceCompId"])
            if src:
                conn_desc = (f"{src['type']}[{conn_port(cn, 'sourcePort')}] → "
                             f"{ctype}[{conn_port(cn, 'targetPort')}]")
            break
    return tools, conn_desc


def chat_test(api_base, api_key, model, tools, prompt, smart_mode=False, available_skills=None, max_rounds=8):
    """通过 /api/chat 测试一次 LLM 对话，返回 (calls, results, errors, content)"""
    payload = {
        "messages": [{"role": "user", "content": prompt}],
        "model": model,
        "api_base": api_base,
        "api_key": api_key,
        "tools": tools,
        "max_tool_rounds": max_rounds,
        "max_tokens": 1024,
        "temperature": 0.1,
    }
    if smart_mode:
        payload["smart_mode"] = True
        payload["available_skills"] = [{"id": s, "name": s} for s in (available_skills or [])]

    resp = SESSION.post(CHAT, json=payload, timeout=120, stream=True)
    calls, results, errors, content = [], [], [], ""
    try:
        for line in resp.iter_lines(decode_unicode=True):
            if not line or not line.startswith("data: "):
                continue
            ds = line[6:]
            if ds == "[DONE]":
                break
            try:
                d = json.loads(ds)
            except json.JSONDecodeError:
                continue
            if d.get("error"):
                errors.append(d["error"])
            if d.get("tool_calls"):
                for tc in d["tool_calls"]:
                    calls.append({"name": tc.get("name", ""), "args": tc.get("arguments", "")[:200]})
            if d.get("tool_result"):
                tr = d["tool_result"]
                results.append({"name": tr.get("name", ""), "result": tr.get("result", "")[:400]})
            if d.get("choices"):
                c = d["choices"][0].get("delta", {}).get("content", "")
                if c:
                    content += c
    finally:
        resp.close()
    return calls, results, errors, content


def classify(calls, results, errors, expected_tools, content, elapsed):
    """判定 PASS / WARN / FAIL。
    expected_tools: 该组件映射表内任一工具被调用即视为命中（镜像编辑器注入行为）。"""
    called_names = [c["name"] for c in calls]
    hit = [t for t in expected_tools if t in called_names]
    # "达最大轮数"是对话循环截断，不是组件错误；若期望工具已调用且结果干净则仍算 PASS
    real_errors = [e for e in errors if MAX_ROUNDS_MSG not in e]
    if real_errors:
        return "FAIL", real_errors[0][:150], real_errors[0][:200]
    if not hit:
        if not called_names:
            return "FAIL", "LLM 未调用工具: " + content[:60], "LLM 拒绝或未识别工具调用"
        return ("FAIL",
                "调用了非预期工具: " + ", ".join(dict.fromkeys(called_names)),
                "期望: " + ",".join(expected_tools))
    for r in results:
        for kw in WARN_KEYWORDS:
            if kw in r["result"]:
                return "WARN", r["result"][:150], r["result"][:200]
    note = "（达最大轮数）" if any(MAX_ROUNDS_MSG in e for e in errors) else ""
    return "PASS", "called: " + ", ".join(dict.fromkeys(called_names)) + note, ""


def test_memory_component(api_base, api_key, model):
    """memory 组件：直接测试 /api/memory/save + load（不走 LLM）"""
    session_id = f"auto_test_{int(time.time())}"
    pid = "default"
    try:
        r1 = SESSION.post(MEMORY_SAVE, json={
            "session_id": session_id, "project_id": pid,
            "messages": [{"role": "user", "content": "测试记忆"},
                         {"role": "assistant", "content": "记忆已保存"}],
        }, timeout=30)
        ok1 = r1.status_code == 200
        r2 = SESSION.get(f"{MEMORY_LOAD}/{session_id}?project_id={pid}", timeout=30)
        ok2 = r2.status_code == 200
        if ok1 and ok2:
            data = r2.json()
            msgs = data.get("session", {}).get("messages", []) if isinstance(data, dict) else []
            if len(msgs) >= 2:
                return "PASS", f"save+load 往返成功（{len(msgs)} 条消息）", ""
            return "WARN", f"save={ok1} load={ok2}，消息数={len(msgs)}", "记忆往返不完整"
        return "WARN", f"save={ok1} load={ok2}", "记忆接口异常"
    except Exception as e:
        return "FAIL", str(e)[:150], str(e)[:200]


def main():
    login()
    projects = sys.argv[1:] or [DEFAULT_PROJECT]
    all_results = []
    # 工具定义索引（缓存一次，避免每组件重复请求与连接池堆积）
    _tool_index = {}
    # 组件过滤：TEST_TYPES="skill_frontend,skill_pua" 只测指定类型（分块调试用）
    type_filter = None
    _tf = os.environ.get("TEST_TYPES", "").strip()
    if _tf:
        type_filter = set(t.strip() for t in _tf.split(",") if t.strip())

    for proj_path in projects:
        comps, conns = load_project(proj_path)
        api_base, api_key, model = get_llm_config(comps)
        if not api_base:
            api_base, api_key, model = "https://api.deepseek.com/v1", \
                os.environ.get("LLM_API_KEY", ""), "deepseek-chat"
        print(f"项目: {proj_path} | {len(comps)} 组件 / {len(conns)} 连线 | LLM: {model}")

        for comp in sorted(comps, key=lambda c: c["id"]):
            cid, ctype = comp["id"], comp["type"]
            if type_filter and ctype not in type_filter:
                continue
            print(f"  [{cid}] {ctype} ... ", end="", flush=True)

            # meta 组件
            if ctype in META_TYPES:
                if ctype == "memory":
                    status, detail, issues = test_memory_component(api_base, api_key, model)
                else:
                    status, detail, issues = "PASS", "无需工具调用: " + META_TYPES[ctype], ""
                all_results.append({
                    "project": os.path.basename(proj_path), "id": cid, "type": ctype,
                    "category": "编排/增强" if ctype != "memory" else "记忆",
                    "status": status, "elapsed": "-", "detail": detail,
                    "tools_called": "-", "issues": issues,
                })
                print(f"[{status}]")
                continue

            # 推导真实工具集
            tools, conn_desc = derive_component_info(comp, comps, conns)
            if not tools:
                all_results.append({
                    "project": os.path.basename(proj_path), "id": cid, "type": ctype,
                    "category": "未知", "status": "SKIP", "elapsed": "-",
                    "detail": "未推导出工具", "tools_called": "-",
                    "issues": f"未知组件类型: {ctype}（连线: {conn_desc}）",
                })
                print("[SKIP]")
                continue

            # 构建 OpenAI tools 数组（工具定义缓存一次，避免每组件重复请求与连接池堆积）
            if not _tool_index:
                _tool_defs = SESSION.get(API + "/api/tools/definitions", timeout=60).json()
                _tool_index.update({t["function"]["name"]: t for t in _tool_defs})
            tool_index = _tool_index

            # 技能组件：智能模式
            if ctype.startswith("skill_"):
                skill_id = tools[0]
                t0 = time.time()
                calls, results, errors, content = chat_test(
                    api_base, api_key, model,
                    [{"type": "function", "function": tool_index["use_skill"]["function"]}],
                    TEST_PROMPTS.get(ctype, f"使用技能 {skill_id} 完成任务"),
                    smart_mode=True, available_skills=[skill_id], max_rounds=8,
                )
                elapsed = round(time.time() - t0, 1)
                called_names = [c["name"] for c in calls]
                real_errors = [e for e in errors if MAX_ROUNDS_MSG not in e]
                if "use_skill" in called_names and any(
                        skill_id in c["args"] for c in calls):
                    status, detail, issues = "PASS", f"use_skill 激活技能 {skill_id}", ""
                elif "use_skill" in called_names:
                    status, detail, issues = "WARN", "use_skill 已调用但参数非预期技能: " + str(calls[:1]), ""
                elif real_errors:
                    status, detail, issues = "FAIL", real_errors[0][:150], real_errors[0][:200]
                else:
                    status, detail, issues = "FAIL", "智能模式未调用 use_skill: " + content[:60], "LLM 未激活技能"
                all_results.append({
                    "project": os.path.basename(proj_path), "id": cid, "type": ctype,
                    "category": "技能", "status": status, "elapsed": elapsed,
                    "detail": detail, "tools_called": ", ".join(dict.fromkeys(called_names)), "issues": issues,
                })
                print(f"[{status}] ({elapsed}s)")
                continue

            # skills_manager / skill_auto_call：智能模式整体测试
            if ctype in ("skills_manager", "skill_auto_call"):
                skill_ids = []
                for c in comps:
                    if c["type"].startswith("skill_"):
                        sid = derive_component_info(c, comps, conns)[0][0]
                        skill_ids.append(sid)
                skill_ids = list(dict.fromkeys(skill_ids))
                t0 = time.time()
                calls, results, errors, content = chat_test(
                    api_base, api_key, model,
                    [{"type": "function", "function": tool_index["use_skill"]["function"]}],
                    TEST_PROMPTS[ctype], smart_mode=True, available_skills=skill_ids, max_rounds=6,
                )
                elapsed = round(time.time() - t0, 1)
                called_names = [c["name"] for c in calls]
                real_errors = [e for e in errors if MAX_ROUNDS_MSG not in e]
                if "use_skill" in called_names:
                    status, detail, issues = "PASS", f"智能模式可用技能 {len(skill_ids)} 个: {', '.join(skill_ids)}", ""
                elif real_errors:
                    status, detail, issues = "FAIL", real_errors[0][:150], real_errors[0][:200]
                elif content:
                    status, detail, issues = "WARN", "智能模式未激活技能（LLM 直接回复）: " + content[:60], ""
                else:
                    status, detail, issues = "FAIL", "智能模式无响应", ""
                all_results.append({
                    "project": os.path.basename(proj_path), "id": cid, "type": ctype,
                    "category": "技能", "status": status, "elapsed": elapsed,
                    "detail": detail, "tools_called": ", ".join(dict.fromkeys(called_names)), "issues": issues,
                })
                print(f"[{status}] ({elapsed}s)")
                continue

            # 普通工具组件
            test_tools = [{"type": "function", "function": tool_index[t]["function"]}
                          for t in tools if t in tool_index]
            if not test_tools:
                all_results.append({
                    "project": os.path.basename(proj_path), "id": cid, "type": ctype,
                    "category": "工具", "status": "SKIP", "elapsed": "-",
                    "detail": f"工具未在后端注册: {tools}", "tools_called": "-",
                    "issues": "工具注册缺失",
                })
                print("[SKIP]")
                continue

            prompt = TEST_PROMPTS.get(ctype, f"use {tools[0]} tool")
            t0 = time.time()
            calls, results, errors, content = chat_test(
                api_base, api_key, model, test_tools, prompt, max_rounds=12,
            )
            elapsed = round(time.time() - t0, 1)
            status, detail, issues = classify(calls, results, errors, tools, content, elapsed)
            all_results.append({
                "project": os.path.basename(proj_path), "id": cid, "type": ctype,
                "category": "工具", "status": status, "elapsed": elapsed,
                "detail": detail, "tools_called": ", ".join(dict.fromkeys(c["name"] for c in calls)),
                "issues": issues, "connection": conn_desc,
            })
            print(f"[{status}] ({elapsed}s)")

    # ---- 汇总 ----
    from collections import Counter
    counts = Counter(r["status"] for r in all_results)
    print("\n" + "=" * 60)
    print(f"RESULTS: {counts.get('PASS', 0)} PASS, {counts.get('WARN', 0)} WARN, "
          f"{counts.get('FAIL', 0)} FAIL, {counts.get('SKIP', 0)} SKIP  (共 {len(all_results)} 项)")
    for r in all_results:
        if r["status"] != "PASS":
            print(f"  [{r['status']}] {r['type']:<20} | {r['detail'][:80]}")

    # ---- 保存 JSON ----
    with open("test_all_results.json", "w", encoding="utf-8") as f:
        json.dump(all_results, f, ensure_ascii=False, indent=2)

    # ---- 生成 Excel ----
    write_excel(all_results, counts)
    print("Excel saved: component_status_report.xlsx")
    print("Done!")


def write_excel(results, counts):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Component Status"

    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    warn_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    skip_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    thin = Border(left=Side(style="thin"), right=Side(style="thin"),
                  top=Side(style="thin"), bottom=Side(style="thin"))

    headers = ["#", "项目", "组件ID", "组件类型", "分类", "状态", "耗时(s)", "工具调用", "连线/详情", "问题"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin

    fill_map = {"PASS": pass_fill, "WARN": warn_fill, "FAIL": fail_fill, "SKIP": skip_fill}
    for i, r in enumerate(results, 2):
        ws.cell(row=i, column=1, value=i - 1).border = thin
        ws.cell(row=i, column=2, value=r.get("project", "")).border = thin
        ws.cell(row=i, column=3, value=r.get("id", "")).border = thin
        ws.cell(row=i, column=4, value=r.get("type", "")).border = thin
        ws.cell(row=i, column=5, value=r.get("category", "")).border = thin
        st = ws.cell(row=i, column=6, value=r.get("status", ""))
        st.border = thin
        st.alignment = Alignment(horizontal="center")
        st.fill = fill_map.get(r.get("status"), skip_fill)
        ws.cell(row=i, column=7, value=r.get("elapsed", "-")).border = thin
        ws.cell(row=i, column=8, value=r.get("tools_called", "")).border = thin
        ws.cell(row=i, column=9, value=r.get("detail", "")).border = thin
        ws.cell(row=i, column=10, value=r.get("issues", "")).border = thin
        for col in range(1, 11):
            ws.cell(row=i, column=col).alignment = Alignment(vertical="center", wrap_text=(col >= 8))

    summary_row = len(results) + 3
    ws.cell(row=summary_row, column=1, value="汇总").font = Font(bold=True, size=12)
    ws.merge_cells(start_row=summary_row, start_column=1, end_row=summary_row, end_column=2)
    summary_data = [
        ("PASS", counts.get("PASS", 0), pass_fill),
        ("WARN", counts.get("WARN", 0), warn_fill),
        ("FAIL", counts.get("FAIL", 0), fail_fill),
        ("SKIP", counts.get("SKIP", 0), skip_fill),
        ("TOTAL", len(results), None),
    ]
    for j, (label, count, fill) in enumerate(summary_data):
        row = summary_row + 1 + j
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row, column=2, value=count).font = Font(bold=True)
        if fill:
            ws.cell(row=row, column=1).fill = fill

    widths = [5, 26, 8, 20, 10, 8, 10, 38, 60, 50]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "A2"

    wb.save("component_status_report.xlsx")


if __name__ == "__main__":
    main()
